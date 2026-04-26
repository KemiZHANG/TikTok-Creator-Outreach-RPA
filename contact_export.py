import re
import sqlite3
from pathlib import Path
from typing import Iterable, List

from openpyxl import Workbook, load_workbook

from utils import BASE_DIR, log


CONTACTS_XLSX = BASE_DIR / "leads_contacts.xlsx"
CONTACTS_DB = BASE_DIR / "contact_leads.db"
HEADERS = ["达人昵称", "邮箱", "商务电话"]

EMAIL_PATTERN = re.compile(
    r"\b[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}\b",
    re.IGNORECASE,
)

PHONE_CONTEXT_KEYWORDS = [
    "inquiries",
    "inquiry",
    "viber",
    "whatsapp",
    "whats app",
    "wa",
    "business",
    "biz",
    "collab",
    "collaboration",
    "collaborate",
    "partnership",
    "contact",
    "phone",
    "tel",
    "telegram",
    "line",
    "dm",
    "call",
    "text",
    "sms",
    "message",
]

PHONE_PATTERN = re.compile(r"(?:(?:\+|00)\s*)?\d[\d\s().\-]{5,}\d")


def init_contact_db() -> None:
    with sqlite3.connect(CONTACTS_DB) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS contact_leads (
                creator_name TEXT,
                email TEXT DEFAULT '',
                phone TEXT DEFAULT '',
                created_time TEXT DEFAULT (datetime('now', 'localtime')),
                UNIQUE(email, phone)
            )
            """
        )
        conn.commit()
    log(f"联系方式数据库已就绪: {CONTACTS_DB}")


def _normalize_text_for_email(text: str) -> str:
    normalized = text or ""
    normalized = normalized.replace("＠", "@")
    normalized = normalized.replace(" gmaii.", " gmail.")
    normalized = normalized.replace("@gmaii.", "@gmail.")
    normalized = normalized.replace("@gmai1.", "@gmail.")
    normalized = normalized.replace("@gma1l.", "@gmail.")
    normalized = re.sub(r"@([a-zA-Z]+)\s*[,，]\s*([a-zA-Z]{2,})", r"@\1.\2", normalized)
    normalized = re.sub(r"@([a-zA-Z]+)\s+([a-zA-Z]{2,})", r"@\1.\2", normalized)
    return normalized


def _repair_common_gmail_ocr_errors(email: str) -> str:
    local, sep, domain = email.partition("@")
    if not sep:
        return email

    domain_lower = domain.lower()
    domain_repairs = {
        "gmaii.com": "gmail.com",
        "gmai1.com": "gmail.com",
        "gma1l.com": "gmail.com",
        "gmali.com": "gmail.com",
        "gmall.com": "gmail.com",
        "gmail.con": "gmail.com",
        "gmail.corn": "gmail.com",
    }
    repaired_domain = domain_repairs.get(domain_lower, domain_lower)
    return f"{local.lower()}@{repaired_domain}"


def extract_emails(text: str) -> List[str]:
    normalized = _normalize_text_for_email(text)
    emails = []
    seen = set()
    for match in EMAIL_PATTERN.findall(normalized):
        email = match.strip(" .,:;，。；）)]}").lower()
        email = _repair_common_gmail_ocr_errors(email)
        if email and email not in seen:
            emails.append(email)
            seen.add(email)
    return emails


def _digits_only(value: str) -> str:
    return re.sub(r"\D", "", value or "")


def _normalize_phone(value: str) -> str:
    value = (value or "").strip()
    has_plus = value.startswith("+")
    digits = _digits_only(value)
    if value.startswith("00") and len(digits) > 2:
        return "+" + digits[2:]
    if has_plus:
        return "+" + digits
    return digits


def _is_plausible_business_phone(value: str) -> bool:
    digits = _digits_only(value)
    if len(digits) < 7 or len(digits) > 15:
        return False
    if len(set(digits)) <= 2:
        return False
    if len(digits) == 4 and digits.startswith(("19", "20")):
        return False
    return True


def extract_business_phones(text: str) -> List[str]:
    lines = [line.strip() for line in (text or "").splitlines() if line.strip()]
    phones = []
    seen = set()

    for index, line in enumerate(lines):
        lower_line = line.lower()
        has_context = any(keyword in lower_line for keyword in PHONE_CONTEXT_KEYWORDS)
        if not has_context:
            continue

        candidate_lines = [line]
        if not PHONE_PATTERN.search(line) and index + 1 < len(lines):
            candidate_lines.append(lines[index + 1])

        for candidate_line in candidate_lines:
            for raw_phone in PHONE_PATTERN.findall(candidate_line):
                normalized_phone = _normalize_phone(raw_phone)
                if not _is_plausible_business_phone(normalized_phone):
                    continue
                if normalized_phone not in seen:
                    phones.append(normalized_phone)
                    seen.add(normalized_phone)

    return phones


def extract_contacts(text: str) -> tuple[List[str], List[str]]:
    return extract_emails(text), extract_business_phones(text)


def _ensure_workbook(path: Path):
    if path.exists():
        workbook = load_workbook(path)
        sheet = workbook.active
        if sheet.max_row == 0:
            sheet.append(HEADERS)
        return workbook, sheet

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "contacts"
    sheet.append(HEADERS)
    workbook.save(path)
    return workbook, sheet


def _contact_exists(email: str, phone: str) -> bool:
    with sqlite3.connect(CONTACTS_DB) as conn:
        cursor = conn.execute(
            "SELECT 1 FROM contact_leads WHERE email = ? AND phone = ? LIMIT 1",
            (email, phone),
        )
        return cursor.fetchone() is not None


def _insert_contact_to_db(creator_name: str, email: str, phone: str) -> bool:
    try:
        with sqlite3.connect(CONTACTS_DB) as conn:
            conn.execute(
                """
                INSERT INTO contact_leads (creator_name, email, phone)
                VALUES (?, ?, ?)
                """,
                (creator_name, email, phone),
            )
            conn.commit()
        return True
    except sqlite3.IntegrityError:
        return False


def append_contact_if_new(creator_name: str, emails: Iterable[str], phones: Iterable[str]) -> bool:
    email_list = list(emails)
    phone_list = list(phones)
    creator_name = (creator_name or "").strip()

    if not email_list and not phone_list:
        log(f"未提取到邮箱或商务电话，不写入联系方式数据库和 Excel: {creator_name}")
        return False

    email_text = "; ".join(email_list)
    phone_text = "; ".join(phone_list)

    if _contact_exists(email_text, phone_text):
        log(f"联系方式数据库中已存在相同记录，跳过 Excel 写入: {(creator_name, email_text, phone_text)}")
        return False

    if not _insert_contact_to_db(creator_name, email_text, phone_text):
        log(f"联系方式数据库写入时发现重复，跳过 Excel 写入: {(creator_name, email_text, phone_text)}")
        return False

    workbook, sheet = _ensure_workbook(CONTACTS_XLSX)
    sheet.append([creator_name, email_text, phone_text])
    workbook.save(CONTACTS_XLSX)
    log(f"已写入联系方式数据库和 Excel: {CONTACTS_XLSX} | {(creator_name, email_text, phone_text)}")
    return True
